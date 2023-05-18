import pandas as pd
import openpyxl as xl
from geocoder import extract_cords


def address_table(excel):
    excel = r"C:\Users\User\Documents\Vet_facilities.xlsx"
    wb = xl.load_workbook(excel)
    sheet = wb['Append1']
    add_list = []
    for row in range(1, sheet.max_row + 1):
        contact = sheet.cell(row, 5)
        address = sheet.cell(row, 4)
        facility_type = sheet.cell(row, 3)
        name = sheet.cell(row, 2)
        name_co_ords = extract_cords(contact.value, address.value, name.value)
        info = name.value, facility_type.value, address.value, contact.value, name_co_ords[0], name_co_ords[1], \
               name_co_ords[2], name_co_ords[3], name_co_ords[4], name_co_ords[5]
        add_list.append(info)
        df = pd.DataFrame(add_list, columns=['Facility Name', 'Facility Type', 'Facility Address', 'Facility Contact',
                                             'Name Lat', 'Name Long', 'Address Lat', 'Address Long', 'Contact Lat',
                                             'Contact Long'])
        writer = pd.ExcelWriter('vet_co-ordinates.xlsx', engine='xlsxwriter')
        df.to_excel(writer, sheet_name='addresses', index=False)
        writer.close()
        print(info)


print(address_table(r"C:\Users\User\Documents\Vet_facilities.xlsx"))
